"""
Export and Reporting Functions for Steel Cutting System
鋼板切断システム用エクスポート・レポート機能

Generates PDF work instructions, Excel reports, and data exports
PDF作業指示書、Excelレポート、データエクスポートを生成
"""

import os
import io
import logging
import json
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from dataclasses import asdict
import base64

# Import required libraries (will be installed)
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.platypus.flowables import Image
    from reportlab.pdfgen import canvas
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab not available. PDF generation will be disabled.")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("OpenPyXL not available. Excel export will be disabled.")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logging.warning("Pandas not available. Advanced data processing will be limited.")

from cutting.instruction import WorkInstruction, CuttingInstruction
from cutting.quality import QualityCheckpoint, QualityManager
from core.models import PlacementResult, Panel


class ExportFormat:
    """Export format constants"""
    PDF = "pdf"
    EXCEL = "xlsx"
    CSV = "csv"
    JSON = "json"
    XML = "xml"


class ReportType:
    """Report type constants"""
    WORK_INSTRUCTION = "work_instruction"
    QUALITY_PLAN = "quality_plan"
    EFFICIENCY_REPORT = "efficiency_report"
    CUTTING_SUMMARY = "cutting_summary"
    MATERIAL_USAGE = "material_usage"


class DocumentExporter:
    """
    Document export and report generation system
    文書エクスポート・レポート生成システム
    """

    def __init__(self):
        self.logger = logging.getLogger(__name__)

        # Document styling
        self.styles = self._initialize_styles()

        # Template paths
        self.template_dir = "templates"
        self.output_dir = "exports"

        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)

    def export_work_instruction_pdf(
        self,
        work_instruction: WorkInstruction,
        quality_checkpoints: Optional[List[QualityCheckpoint]] = None,
        include_diagrams: bool = True
    ) -> str:
        """
        Export work instruction as PDF document
        作業指示をPDF文書としてエクスポート
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF export")

        filename = f"work_instruction_{work_instruction.sheet_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.output_dir, filename)

        self.logger.info(f"Generating PDF work instruction: {filepath}")

        # Create PDF document
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )

        # Build document content
        story = []

        # Title page
        story.extend(self._create_title_page(work_instruction))
        story.append(PageBreak())

        # Summary section
        story.extend(self._create_summary_section(work_instruction))
        story.append(Spacer(1, 12))

        # Safety and setup section
        story.extend(self._create_safety_section(work_instruction))
        story.append(Spacer(1, 12))

        # Cutting sequence section
        story.extend(self._create_cutting_sequence_section(work_instruction))
        story.append(PageBreak())

        # Quality checkpoints section
        if quality_checkpoints:
            story.extend(self._create_quality_section(quality_checkpoints))
            story.append(PageBreak())

        # Machine settings section
        story.extend(self._create_machine_settings_section(work_instruction))

        # Build PDF
        doc.build(story)

        self.logger.info(f"PDF work instruction generated: {filepath}")
        return filepath

    def export_efficiency_report_excel(
        self,
        placement_results: List[PlacementResult],
        panels: List[Panel]
    ) -> str:
        """
        Export efficiency analysis as Excel workbook
        効率分析をExcelワークブックとしてエクスポート
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("OpenPyXL is required for Excel export")

        filename = f"efficiency_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        self.logger.info(f"Generating Excel efficiency report: {filepath}")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create summary sheet
        self._create_summary_sheet(wb, placement_results, panels)

        # Create detailed results sheet
        self._create_detailed_results_sheet(wb, placement_results)

        # Create material analysis sheet
        self._create_material_analysis_sheet(wb, placement_results, panels)

        # Create panel utilization sheet
        self._create_panel_utilization_sheet(wb, placement_results, panels)

        # Save workbook
        wb.save(filepath)

        self.logger.info(f"Excel efficiency report generated: {filepath}")
        return filepath

    def export_cutting_summary_json(
        self,
        work_instructions: List[WorkInstruction],
        placement_results: List[PlacementResult]
    ) -> str:
        """
        Export cutting summary as JSON data
        切断サマリーをJSONデータとしてエクスポート
        """
        filename = f"cutting_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(self.output_dir, filename)

        self.logger.info(f"Generating JSON cutting summary: {filepath}")

        # Compile summary data
        summary_data = {
            'export_info': {
                'generated_at': datetime.now().isoformat(),
                'generator': 'Steel Cutting Optimization System',
                'version': '1.0.0'
            },
            'overview': {
                'total_sheets': len(placement_results),
                'total_work_instructions': len(work_instructions),
                'total_cutting_time': sum(wi.estimated_total_time for wi in work_instructions),
                'average_efficiency': sum(pr.efficiency for pr in placement_results) / len(placement_results) if placement_results else 0
            },
            'material_breakdown': self._compile_material_breakdown(placement_results),
            'work_instructions': [self._serialize_work_instruction(wi) for wi in work_instructions],
            'placement_results': [self._serialize_placement_result(pr) for pr in placement_results]
        }

        # Write JSON file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(summary_data, f, indent=2, ensure_ascii=False)

        self.logger.info(f"JSON cutting summary generated: {filepath}")
        return filepath

    def _initialize_styles(self) -> Dict[str, Any]:
        """Initialize document styles"""
        styles = getSampleStyleSheet()

        # Custom styles
        custom_styles = {
            'title': ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=24,
                textColor=colors.darkblue,
                alignment=TA_CENTER,
                spaceAfter=20
            ),
            'heading1': ParagraphStyle(
                'CustomHeading1',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.darkblue,
                spaceBefore=12,
                spaceAfter=6
            ),
            'heading2': ParagraphStyle(
                'CustomHeading2',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.darkgreen,
                spaceBefore=10,
                spaceAfter=4
            ),
            'body_japanese': ParagraphStyle(
                'BodyJapanese',
                parent=styles['Normal'],
                fontSize=10,
                fontName='HeiseiKakuGo-W5',
                leading=14
            ),
            'table_header': ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.white,
                alignment=TA_CENTER
            ),
            'safety_warning': ParagraphStyle(
                'SafetyWarning',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.red,
                leftIndent=10,
                rightIndent=10,
                spaceBefore=6,
                spaceAfter=6
            )
        }

        return {**styles, **custom_styles}

    def _create_title_page(self, work_instruction: WorkInstruction) -> List[Any]:
        """Create title page elements"""
        elements = []

        # Main title
        title_text = f"作業指示書 / Work Instruction<br/>Sheet ID: {work_instruction.sheet_id}"
        elements.append(Paragraph(title_text, self.styles['title']))
        elements.append(Spacer(1, 30))

        # Basic information table
        info_data = [
            ['項目 / Item', '内容 / Details'],
            ['材質 / Material', work_instruction.material_type],
            ['シートサイズ / Sheet Size', f"{work_instruction.sheet_dimensions[0]} × {work_instruction.sheet_dimensions[1]} mm"],
            ['総ステップ数 / Total Steps', str(work_instruction.total_steps)],
            ['予想作業時間 / Estimated Time', f"{work_instruction.estimated_total_time:.1f} minutes"],
            ['ケルフ幅 / Kerf Width', f"{work_instruction.kerf_width} mm"],
            ['生成日時 / Generated At', work_instruction.generated_at.strftime('%Y-%m-%d %H:%M:%S')],
            ['生成者 / Generated By', work_instruction.generated_by]
        ]

        info_table = Table(info_data, colWidths=[80*mm, 100*mm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(info_table)
        elements.append(Spacer(1, 30))

        # Complexity assessment
        complexity_score = work_instruction.complexity_score
        complexity_text = "低 / Low" if complexity_score < 0.3 else "中 / Medium" if complexity_score < 0.7 else "高 / High"

        complexity_para = Paragraph(
            f"<b>作業複雑度 / Work Complexity:</b> {complexity_text} ({complexity_score:.2f})",
            self.styles['Normal']
        )
        elements.append(complexity_para)

        return elements

    def _create_summary_section(self, work_instruction: WorkInstruction) -> List[Any]:
        """Create summary section"""
        elements = []

        elements.append(Paragraph("概要 / Summary", self.styles['heading1']))

        # Key metrics
        metrics_data = [
            ['指標 / Metric', '値 / Value'],
            ['総切断長 / Total Cut Length', f"{work_instruction.total_cut_length:.1f} mm"],
            ['作業ステップ数 / Work Steps', str(work_instruction.total_steps)],
            ['予想時間 / Estimated Time', f"{work_instruction.estimated_total_time:.1f} minutes"],
            ['平均ステップ時間 / Avg Step Time', f"{work_instruction.estimated_total_time/work_instruction.total_steps:.1f} min"],
        ]

        metrics_table = Table(metrics_data, colWidths=[90*mm, 70*mm])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))

        elements.append(metrics_table)

        return elements

    def _create_safety_section(self, work_instruction: WorkInstruction) -> List[Any]:
        """Create safety section"""
        elements = []

        elements.append(Paragraph("安全注意事項 / Safety Notes", self.styles['heading1']))

        for i, note in enumerate(work_instruction.safety_notes, 1):
            elements.append(Paragraph(f"{i}. {note}", self.styles['safety_warning']))

        return elements

    def _create_cutting_sequence_section(self, work_instruction: WorkInstruction) -> List[Any]:
        """Create cutting sequence section"""
        elements = []

        elements.append(Paragraph("切断手順 / Cutting Sequence", self.styles['heading1']))

        # Create cutting sequence table
        sequence_data = [['Step', 'Type / タイプ', 'Start / 開始点', 'End / 終了点', 'Length / 長さ', 'Time / 時間']]

        for instruction in work_instruction.cutting_sequence:
            cut_type_jp = "横切断" if instruction.cut_type.value == "horizontal" else "縦切断"
            start_point = f"({instruction.start_point[0]:.1f}, {instruction.start_point[1]:.1f})"
            end_point = f"({instruction.end_point[0]:.1f}, {instruction.end_point[1]:.1f})"

            sequence_data.append([
                str(instruction.step_number),
                f"{instruction.cut_type.value}<br/>{cut_type_jp}",
                start_point,
                end_point,
                f"{instruction.dimension:.1f} mm",
                f"{instruction.estimated_time:.1f} min"
            ])

        sequence_table = Table(sequence_data, colWidths=[15*mm, 30*mm, 30*mm, 30*mm, 25*mm, 20*mm])
        sequence_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(sequence_table)

        return elements

    def _create_quality_section(self, quality_checkpoints: List[QualityCheckpoint]) -> List[Any]:
        """Create quality checkpoints section"""
        elements = []

        elements.append(Paragraph("品質チェックポイント / Quality Checkpoints", self.styles['heading1']))

        # Create quality checkpoints table
        quality_data = [['ID', 'Step / ステップ', 'Type / タイプ', 'Description / 説明', 'Time / 時間']]

        for checkpoint in quality_checkpoints:
            quality_data.append([
                checkpoint.checkpoint_id,
                str(checkpoint.step_number),
                checkpoint.checkpoint_type.value,
                f"{checkpoint.description}<br/>{checkpoint.japanese_description}",
                f"{checkpoint.estimated_time:.1f} min"
            ])

        quality_table = Table(quality_data, colWidths=[20*mm, 15*mm, 25*mm, 80*mm, 20*mm])
        quality_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(quality_table)

        return elements

    def _create_machine_settings_section(self, work_instruction: WorkInstruction) -> List[Any]:
        """Create machine settings section"""
        elements = []

        elements.append(Paragraph("機械設定 / Machine Settings", self.styles['heading1']))

        settings_data = [['設定項目 / Setting', '値 / Value']]

        for key, value in work_instruction.machine_settings.items():
            key_display = key.replace('_', ' ').title()
            settings_data.append([key_display, str(value)])

        settings_table = Table(settings_data, colWidths=[80*mm, 80*mm])
        settings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))

        elements.append(settings_table)

        return elements

    def _create_summary_sheet(self, wb: Workbook, placement_results: List[PlacementResult], panels: List[Panel]):
        """Create Excel summary sheet"""
        ws = wb.create_sheet("Summary / 概要")

        # Headers
        ws['A1'] = "鋼板切断最適化サマリー / Steel Cutting Optimization Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        # Basic statistics
        row = 3
        stats = [
            ("総シート数 / Total Sheets", len(placement_results)),
            ("総パネル数 / Total Panels", len(panels)),
            ("平均効率 / Average Efficiency", f"{sum(pr.efficiency for pr in placement_results) / len(placement_results):.1%}" if placement_results else "0%"),
            ("総廃材面積 / Total Waste Area", f"{sum(pr.waste_area for pr in placement_results):.0f} mm²"),
            ("総切断長 / Total Cut Length", f"{sum(pr.cut_length for pr in placement_results):.0f} mm"),
        ]

        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

    def _create_detailed_results_sheet(self, wb: Workbook, placement_results: List[PlacementResult]):
        """Create detailed results sheet"""
        ws = wb.create_sheet("Detailed Results / 詳細結果")

        headers = ["Sheet ID", "Material", "Efficiency", "Waste Area (mm²)", "Cut Length (mm)", "Panels Placed", "Algorithm"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        for row, result in enumerate(placement_results, 2):
            ws.cell(row=row, column=1, value=result.sheet_id)
            ws.cell(row=row, column=2, value=result.material_block)
            ws.cell(row=row, column=3, value=f"{result.efficiency:.1%}")
            ws.cell(row=row, column=4, value=result.waste_area)
            ws.cell(row=row, column=5, value=result.cut_length)
            ws.cell(row=row, column=6, value=len(result.panels))
            ws.cell(row=row, column=7, value=result.algorithm)

    def _create_material_analysis_sheet(self, wb: Workbook, placement_results: List[PlacementResult], panels: List[Panel]):
        """Create material analysis sheet"""
        ws = wb.create_sheet("Material Analysis / 材質分析")

        # Material breakdown
        material_stats = {}
        for panel in panels:
            if panel.material not in material_stats:
                material_stats[panel.material] = {'count': 0, 'area': 0}
            material_stats[panel.material]['count'] += panel.quantity
            material_stats[panel.material]['area'] += panel.area * panel.quantity

        headers = ["Material", "Panel Count", "Total Area (mm²)", "Percentage"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        total_area = sum(stats['area'] for stats in material_stats.values())

        for row, (material, stats) in enumerate(material_stats.items(), 2):
            percentage = (stats['area'] / total_area * 100) if total_area > 0 else 0
            ws.cell(row=row, column=1, value=material)
            ws.cell(row=row, column=2, value=stats['count'])
            ws.cell(row=row, column=3, value=stats['area'])
            ws.cell(row=row, column=4, value=f"{percentage:.1f}%")

    def _create_panel_utilization_sheet(self, wb: Workbook, placement_results: List[PlacementResult], panels: List[Panel]):
        """Create panel utilization sheet"""
        ws = wb.create_sheet("Panel Utilization / パネル利用")

        headers = ["Panel ID", "Width", "Height", "Quantity", "Material", "Placed", "Utilization"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        # Calculate panel utilization
        placed_panels = set()
        for result in placement_results:
            for placed_panel in result.panels:
                placed_panels.add(placed_panel.panel.id)

        for row, panel in enumerate(panels, 2):
            utilization = "100%" if panel.id in placed_panels else "0%"
            ws.cell(row=row, column=1, value=panel.id)
            ws.cell(row=row, column=2, value=panel.width)
            ws.cell(row=row, column=3, value=panel.height)
            ws.cell(row=row, column=4, value=panel.quantity)
            ws.cell(row=row, column=5, value=panel.material)
            ws.cell(row=row, column=6, value="Yes" if panel.id in placed_panels else "No")
            ws.cell(row=row, column=7, value=utilization)

    def _compile_material_breakdown(self, placement_results: List[PlacementResult]) -> Dict[str, Any]:
        """Compile material breakdown statistics"""
        material_stats = {}

        for result in placement_results:
            material = result.material_block
            if material not in material_stats:
                material_stats[material] = {
                    'sheet_count': 0,
                    'total_efficiency': 0,
                    'total_waste': 0,
                    'total_cut_length': 0
                }

            material_stats[material]['sheet_count'] += 1
            material_stats[material]['total_efficiency'] += result.efficiency
            material_stats[material]['total_waste'] += result.waste_area
            material_stats[material]['total_cut_length'] += result.cut_length

        # Calculate averages
        for material, stats in material_stats.items():
            sheet_count = stats['sheet_count']
            stats['average_efficiency'] = stats['total_efficiency'] / sheet_count
            stats['average_waste'] = stats['total_waste'] / sheet_count
            stats['average_cut_length'] = stats['total_cut_length'] / sheet_count

        return material_stats

    def _serialize_work_instruction(self, work_instruction: WorkInstruction) -> Dict[str, Any]:
        """Serialize work instruction for JSON export"""
        return {
            'sheet_id': work_instruction.sheet_id,
            'material_type': work_instruction.material_type,
            'sheet_dimensions': work_instruction.sheet_dimensions,
            'total_steps': work_instruction.total_steps,
            'estimated_total_time': work_instruction.estimated_total_time,
            'kerf_width': work_instruction.kerf_width,
            'generated_at': work_instruction.generated_at.isoformat(),
            'cutting_sequence': [
                {
                    'step_number': inst.step_number,
                    'cut_type': inst.cut_type.value,
                    'start_point': inst.start_point,
                    'end_point': inst.end_point,
                    'dimension': inst.dimension,
                    'estimated_time': inst.estimated_time,
                    'safety_level': inst.safety_level.value
                }
                for inst in work_instruction.cutting_sequence
            ],
            'machine_settings': work_instruction.machine_settings,
            'safety_notes': work_instruction.safety_notes
        }

    def _serialize_placement_result(self, placement_result: PlacementResult) -> Dict[str, Any]:
        """Serialize placement result for JSON export"""
        return {
            'sheet_id': placement_result.sheet_id,
            'material_block': placement_result.material_block,
            'efficiency': placement_result.efficiency,
            'waste_area': placement_result.waste_area,
            'cut_length': placement_result.cut_length,
            'cost': placement_result.cost,
            'algorithm': placement_result.algorithm,
            'processing_time': placement_result.processing_time,
            'timestamp': placement_result.timestamp.isoformat(),
            'panels_placed': len(placement_result.panels),
            'panels': [
                {
                    'panel_id': placed.panel.id,
                    'x': placed.x,
                    'y': placed.y,
                    'width': placed.actual_width,
                    'height': placed.actual_height,
                    'rotated': placed.rotated,
                    'material': placed.panel.material
                }
                for placed in placement_result.panels
            ]
        }


def create_document_exporter() -> DocumentExporter:
    """Create document exporter instance"""
    return DocumentExporter()


# Utility functions for different export formats
def export_to_csv(data: List[Dict[str, Any]], filename: str) -> str:
    """Export data to CSV format"""
    if not PANDAS_AVAILABLE:
        raise ImportError("Pandas is required for CSV export")

    df = pd.DataFrame(data)
    filepath = os.path.join("exports", filename)
    df.to_csv(filepath, index=False, encoding='utf-8-sig')
    return filepath


def check_export_dependencies() -> Dict[str, bool]:
    """Check availability of export dependencies"""
    return {
        'reportlab': REPORTLAB_AVAILABLE,
        'openpyxl': OPENPYXL_AVAILABLE,
        'pandas': PANDAS_AVAILABLE
    }


def install_export_dependencies() -> str:
    """Return pip install command for export dependencies"""
    return "pip install reportlab openpyxl pandas"